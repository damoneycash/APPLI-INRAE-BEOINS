from numpy import *
from pandas import *
from openpyxl import *
import streamlit as st
def saut_ligne(i):
    for i in range(int(i)):
        st.write("")
    return

def main():
    st.write('<img class="Logo-etat" src="https://www.inrae.fr/themes/custom/inrae_socle/public/images/etat_logo.svg" alt="République française" width="138" height="146">',
             '<img class="Logo-site" src="https://www.inrae.fr/themes/custom/inrae_socle/logo.svg" alt="INRAE">',
             unsafe_allow_html=True)
    st.header(":blue[INRAE Besoins individuels]")
    uploaded_file = st.file_uploader("Téléversez un fichier Excel", type=["xlsx", "xls"])
    uploaded_file = 'excel.xlsx'
    if uploaded_file is not None:
        try:
            # Lecture du fichier Excel avec Pandas
            classeur = load_workbook(uploaded_file)
            wb = classeur.sheetnames
            wa = wb[1]
            drap = classeur[wa]

            # Extraction des données pour les pôles
            L = []
            for i in drap["N"]:
                L.append(str(i.value))
            #print(L)
            L.pop(0)
            L.pop(0)
            L.pop(0)
            L.sort()
            Lp = unique([x for x in L if x and x != "None"])

            # Extraction des données pour les agents
            agent = []
            for row in drap.iter_rows(min_row=4, values_only=True):
                nom = row[3]
                prenom = row[4]
                if nom and prenom:
                    nom_prenom = f"{nom} {prenom}"
                    if nom_prenom.strip():
                        agent.append(nom_prenom.strip())
            La = agent
            #print(agent)
            # Extraction des données complètes
            data = []
            for i, row in enumerate(drap.iter_rows(values_only=True), start=1):
                if i == 2:  # Exclure la deuxième ligne (titres)
                    continue
                nom = row[3]
                prenom = row[4]
                if nom and prenom:
                    nom_prenom = f"{nom} {prenom}"
                    data_row = list(row[:14])
                    data_row[3] = nom_prenom
                    del data_row[4]
                    #del data_row[6]
                    data.append(data_row)
            #print(data)
            # Trier les données par ordre alphabétique du nom de famille
            data.sort(key=lambda x: x[3] if x[3] is not None else "")
            #print(Lp)
            #print(data)
            # Organiser les données en fonction des pôles
            data_organized = []
            for i in range(len(Lp)):
                L1 = []
                for j in range(len(data)):
                    if data[j][len(data[j])-1] == Lp[i]:
                        L1.append(data[j])
                data_organized.append(L1)
            selected_pole = st.selectbox("Choisissez un pôle", Lp)

            # Filtrer les données correspondantes au pôle sélectionné
            filtered_data = [row for row in data if row[-1] == selected_pole]

            # Créer le menu déroulant pour choisir un agent parmi le pôle sélectionné
            selected_agent = st.selectbox("Choisissez un agent", [row[3] for row in filtered_data])

            # Filtrer les données correspondantes à l'agent sélectionné
            filtered_data_agent = [row for row in filtered_data if row[3] == selected_agent]

            # Créer le DataFrame avec les données filtrées pour le premier tableau
            df = DataFrame(filtered_data_agent)
    
            header = [cell.value for cell in drap[2]]
            header.pop(4)
            header = header[:len(df.columns)]
            df.columns = header
            st.dataframe(df, use_container_width=False, height=140, hide_index=True)
            st.column_config.TextColumn(width="small")

            saut_ligne(5)

            # Créer le DataFrame avec les données filtrées pour le deuxième tableau
            df3 = DataFrame(filtered_data)
            #header = [cell.value for cell in drap[2]]
            header = header[:len(df3.columns)]
            df3.columns = header
            st.dataframe(df3, use_container_width=True, height=400, hide_index=True)
            st.column_config.TextColumn(width="small")
        except UnboundLocalError:
            # Afficher "Hello World" en cas d'erreur
            st.write("Hello World")
    return
