from openpyxl import load_workbook
from numpy import unique
import streamlit as st
import pandas as pd

def main():
    st.write('<img class="Logo-etat" src="https://www.inrae.fr/themes/custom/inrae_socle/public/images/etat_logo.svg" alt="République française" width="138" height="146">',
             '<img class="Logo-site" src="https://www.inrae.fr/themes/custom/inrae_socle/logo.svg" alt="INRAE">',
             unsafe_allow_html=True)
    st.header(":blue[INRAE Besoins collectifs]")
    uploaded_file = st.file_uploader("Téléversez un fichier Excel", type=["xlsx", "xls"])
    uploaded_file = 'excel.xlsx'
    if uploaded_file is not None:
        try:
            # Lecture du fichier Excel avec Pandas
            classeur = load_workbook(uploaded_file)
            wb = classeur.sheetnames
            wa = wb[2]
            drap2 = classeur[wa]
            header = [cell.value for cell in drap2[2]]
            data_sheet3 = []
            for row in drap2.iter_rows(min_row=3, values_only=True):
                data_sheet3.append(row)

            # Extraction des noms de pôle de la colonne avant-dernière
            poles_sheet3 = []
            for row in data_sheet3:
                pole_value = row[-2]
                if pole_value:
                    poles = pole_value.split(" ")
                    poles = [p for p in poles if p not in ["-", "/", "de", "des", "et", "la"]]
                    poles_sheet3.extend(poles)

            poles_sheet3 = unique([x for x in poles_sheet3 if x])
            
            print(poles_sheet3)
            # Création du menu déroulant pour choisir un pôle (feuille 3)
            selected_pole_sheet3 = st.selectbox("Choisissez un pôle", poles_sheet3)

            # Filtrage des données correspondantes au pôle sélectionné (feuille 3)
            filtered_data_sheet3 = [row for row in data_sheet3 if row[-2] and selected_pole_sheet3 in row[-2]]

            # Renommage des colonnes pour éviter les doublons
            num_columns = len(filtered_data_sheet3[0])
            column_names = ['Colonne_' + str(i) for i in range(num_columns)]
            df_sheet3 = pd.DataFrame(filtered_data_sheet3, columns=column_names)
            df_sheet3.columns = header

            # Affichage du DataFrame (feuille 3)
            st.dataframe(df_sheet3, use_container_width=True, height=800, hide_index=True)
            st.column_config.TextColumn(width="small")

        except UnboundLocalError:
            # Afficher "Hello World" en cas d'erreur
            st.write("Hello World")

    return

main()
